from flask import Blueprint, jsonify, request
import openpyxl
import random
import io

station_names = []
stations = []
stations_api = Blueprint("stations_api", __name__)


def load_station_names():
    with io.open("./data/Stations.txt", encoding='utf-8') as file:
        for line in file:
            station_names.append(line.rstrip())


def load_stations_data():
    stations_data_path = "./data/STATION_COORDS_HACKATON.xlsx"
    wb = openpyxl.load_workbook(stations_data_path)
    sheet = wb.active
    for i in range(2, sheet.max_row):
        station = {
            "id": sheet.cell(row=i, column=1).value,
            "name": random.choice(station_names),
            "latitude": sheet.cell(row=i, column=2).value,
            "longitude": sheet.cell(row=i, column=3).value
        }
        stations.append(station)


def get_stations_by_id(station_id):
    for s in stations:
        if s["id"] == int(station_id):
            return s


def get_stations_data(page, size):
    start = page*size
    end = (page+1)*size
    return stations[start:end]


load_station_names()
load_stations_data()


@stations_api.route("/api/stations")
def get_stations():
    page = request.args.get("page")
    size = request.args.get("size")
    if page:
        page = int(page)
    else:
        page = 0
    if size:
        size = int(size)
    else:
        size = 100
    stations_data = get_stations_data(page, size)
    return jsonify(stations_data), 200


@stations_api.route("/api/stations/<station_id>")
def get_station(station_id):
    station = get_stations_by_id(station_id)
    if station:
        return jsonify(station), 200
    return jsonify({"message": "NotFound"}), 404


@stations_api.route("/api/stations/search")
def search_station():
    latitude = float(request.args.get("latitude"))
    longitude = float(request.args.get("longitude"))
    radius = float(request.args.get("radius"))
    stations_in_radius = []
    for station in stations:
        #  (x — x_0)^2 + (y — y_0)^2 <= R^2
        if station["longitude"] is not None and station["latitude"] is not None and pow((station["latitude"] - latitude), 2) + pow((station["longitude"] - longitude), 2) <= pow(radius, 2):
            stations_in_radius.append(station)
    return jsonify(stations_in_radius), 200
