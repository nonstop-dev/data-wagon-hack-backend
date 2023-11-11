import openpyxl
import threading
from flask import Blueprint, request, jsonify
from api.stations_api import get_stations_by_id

wagons_api = Blueprint("wagons_api", __name__)
wagons = []


def load_wagons_data():
    wagons_data_path = "./data/disl_hackaton.xlsx"
    wb = openpyxl.load_workbook(filename=wagons_data_path, read_only=True, keep_vba=False)
    sheets = wb.sheetnames
    i = 1
    for s_name in sheets:
        sheet = wb[s_name]
        for i in range(2, sheet.max_row):
            train_info_value = sheet.cell(row=i, column=5).value
            train_info = ""
            if train_info_value != "":
                train_info = train_info_value.split("-")
            station_id = sheet.cell(row=i, column=3).value
            station = get_stations_by_id(station_id)
            wagon = {
                "wagonId": sheet.cell(row=i, column=1).value,
                "stationId": station_id,
                "arrivalTime": sheet.cell(row=i, column=2).value,
                "wagonDestination": sheet.cell(row=i, column=4).value,
                "trainDeparturePoint": train_info[0],
                "trainDestinationPoint": train_info[2],
                "trainNumber": train_info[1],
                "latitude": "" if station is None else station["latitude"],
                "longitude": "" if station is None else station["longitude"]
            }
            wagons.append(wagon)


def get_wagons_data(page, size):
    start = page*size
    end = (page+1)*size
    return wagons[start:end]


thr = threading.Thread(target=load_wagons_data, args=(), kwargs={})
thr.start()


@wagons_api.route("/api/wagons")
def get_wagons():
    page = request.args.get("page")
    size = request.args.get("size")
    if page:
        page = int(page)
    else:
        page = 0
    if size:
        size = int(size)
    else:
        size = 100

    wagons_data = get_wagons_data(page, size)

    return jsonify(wagons_data), 200
