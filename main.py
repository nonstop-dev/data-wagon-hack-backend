import openpyxl
import random
import io
import threading

from flask import Flask, request, jsonify

app = Flask(__name__)

wagons = []
station_names = []
stations = []


def load_wagons_data():
    wagons_data_path = "data/disl_hackaton.xlsx"
    wb = openpyxl.load_workbook(filename=wagons_data_path, read_only=True, keep_vba=False)
    sheets = wb.sheetnames
    i = 1
    for s_name in sheets:
        sheet = wb[s_name]
        for i in range(2, sheet.max_row):
            train_info_value = sheet.cell(row=i, column=5).value
            train_info = ""
            if train_info_value != "":
                train_info = train_info_value.split("-")
            wagon = {
                "wagonId": sheet.cell(row=i, column=1).value,
                "stationId": sheet.cell(row=i, column=3).value,
                "arrivalTime": sheet.cell(row=i, column=2).value,
                "wagonDestination": sheet.cell(row=i, column=4).value,
                "trainDeparturePoint": train_info[0],
                "trainDestinationPoint": train_info[2],
                "trainNumber": train_info[1]
            }
            wagons.append(wagon)


def get_wagons_data(page, size):
    start = page*size
    end = (page+1)*size
    return wagons[start:end]


def load_stations_data():
    stations_data_path = "data/STATION_COORDS_HACKATON.xlsx"
    wb = openpyxl.load_workbook(stations_data_path)
    sheet = wb.active
    for i in range(2, sheet.max_row):
        station = {
            "id": sheet.cell(row=i, column=1).value,
            "name": random.choice(station_names),
            "latitude": sheet.cell(row=i, column=2).value,
            "longitude": sheet.cell(row=i, column=3).value
        }
        stations.append(station)


def load_station_names():
    with io.open("data/Stations.txt", encoding='utf-8') as file:
        for line in file:
            station_names.append(line.rstrip())


load_station_names()
load_stations_data()
thr = threading.Thread(target=load_wagons_data, args=(), kwargs={})
thr.start()


@app.route("/stations")
def get_stations():
    return jsonify(stations), 200


@app.route("/wagons")
def get_wagons():
    page = request.args.get("page")
    size = request.args.get("size")
    if page:
        page = int(page)
    else:
        page = 0
    if size:
        size = int(size)
    else:
        size = 100

    wagons_data = wagons[page:size]

    return jsonify(wagons_data), 200


@app.route("/trains/<train_number>")
def get_wagon(train_number):
    pgk_wagons_in_train = []
    train_info = {}
    train_info_raw = {}
    for wagon in wagons:
        if wagon["trainNumber"] == train_number:
            pgk_wagons_in_train.append(wagon["wagonId"])
            train_info_raw["trainDeparturePoint"] = int(wagon["trainDeparturePoint"])
            train_info_raw["trainDestinationPoint"] = int(wagon["trainDestinationPoint"])

    for wagon in wagons:
        if wagon["stationId"] == train_info_raw["trainDeparturePoint"]:
            train_info_raw["departureAt"] = wagon["arrivalTime"]
        if wagon["stationId"] == train_info_raw["trainDestinationPoint"]:
            train_info_raw["arriveAt"] = wagon["arrivalTime"]

    for station in stations:
        if station["id"] == train_info_raw["trainDeparturePoint"]:
            train_info["from"] = station["name"]
        if station["id"] == train_info_raw["trainDestinationPoint"]:
            train_info["to"] = station["name"]

    train_info["arriveAt"] = train_info_raw["arriveAt"] if train_info_raw["arriveAt"] else ""
    train_info["departureAt"] = train_info_raw["departureAt"] if train_info_raw["departureAt"] else ""
    train_info["wagons"] = pgk_wagons_in_train

    return jsonify(train_info), 200


if __name__ == "__main__":
    app.run()
